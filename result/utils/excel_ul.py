import openpyxl
from datetime import datetime
from config.excel import *


def getMark(sheet):
    
    marks = dict()
    absentNumber =  []
    max_row = sheet.max_row

    row = ROLL_NUMBER_START_ROW + 1
    col = ROLL_NUMBER_START_COL + 1
    
    for i in range(row, max_row):
        if(sheet.cell(i, col+1).value == 'AB'):
            absentNumber.append(sheet.cell(i, col).value)
        elif(sheet.cell(i, col+1).value == None):
            break
        else:
            marks[sheet.cell(i, col).value] = sheet.cell(i,col+1).value
            
    return marks, absentNumber


def generateRank(marks_dict, absentNumber):

    marks_dict = dict(sorted(marks_dict.items(), key = lambda x : x[1], reverse=True))
        
    rank = 0
    previousMark = -1
    ranked = dict()
    
    for rollnum, mark in marks_dict.items():
        
        if(mark == previousMark):
            ranked[rollnum] = rank
        else:
            rank += 1
            ranked[rollnum] = rank
            previousMark = mark

    rank += 1
    for rollnum in absentNumber:
         ranked[rollnum] = rank
            
    return ranked
