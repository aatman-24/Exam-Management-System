from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from django.contrib import messages

import openpyxl
from xlsxwriter.workbook import Workbook
from examination.models import Exam
from student.models import Student, AcademicProfile
from django.shortcuts import get_object_or_404

from config.excel import *
from ..models import Result
from ..forms import UploadFileForm, ResultForm
from ..utils import add_data_to_worksheet, getMark, generateRank

from student.views import add_student_subject_marks
from study.views import updateExamRecord
from users.decorator import class_login_required, class_require_authenticated_permisssion

def handle_uploaded_file(f, fname):
    with open(f'store/{fname}.xlsx', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
            
def makeResult(fileName, sheetName='Sheet1'):
    
    wb = openpyxl.load_workbook(fileName)
    ws =  wb[sheetName]

    col = DETAILS_START_COL + 2
    row = DETAILS_START_ROW + 1 

    examId = ws.cell(row = row, column = col).value
    examName = ws.cell(row = row + 1, column = col).value
    examDate = ws.cell(row = row + 2, column = col).value
    std = ws.cell(row = row + 3, column = col).value
    subject = ws.cell(row = row + 4, column = col).value
    totalMarks = ws.cell(row = row + 5, column = col).value
    slug = ws.cell(row = row + 6, column = col).value

    # Get Exam
    exam = get_object_or_404(Exam, id = examId) 
    
    marks, absentNumber = getMark(ws)
    ranked = generateRank(marks, absentNumber)

    for rollnum, mark in marks.items():

        student = get_object_or_404(Student, rollNumber = rollnum, std = std)

        created_result = Result.objects.create(
            exam = exam,
            student = student,
            marks = mark,
            rank = ranked[rollnum],
            absent = (mark == 'AB'),
        )
        created_result.save()
        
        # Add Subject Marks into Student Academic Profile.
        add_student_subject_marks(student, exam.subject, mark) 
        
    # Add Exam into Relative Subject
    updateExamRecord(exam)
            
@class_require_authenticated_permisssion('result.publish_result')
class UploadResultFile(View):

    def get(self, request, exam_slug):
        exam = get_object_or_404(Exam, slug = exam_slug)
        form = UploadFileForm()
        return render(request, 'result/upload.html', {'exam':exam, 'form':form})

    def post(self, request, exam_slug):
        form = UploadFileForm(request.POST, request.FILES)
        exam = get_object_or_404(Exam, slug = exam_slug)
        if form.is_valid():
            file = request.FILES['file']
            # fname = str(exam.id) + "_" + exam.examName + "_" + str(exam.examDate)
            # handle_uploaded_file(request.FILES['file'], fname)
            # messages.success(request, "File is uploaded")
            
            makeResult(file)

            #TODO : WHAT I GO FORM HERE ?
            exam.resultPublished = True
            return HttpResponse("Success")
        return render(request, 'result/okay.html', {'form': form})

@class_login_required
class GetResult(View):

    model = Result

    def get(self, request, result_slug):
        result = get_object_or_404(self.model, slug = result_slug)
        return render(request, 'result/result_detail.html', {'exam' : result.exam, 'result':result})
@class_require_authenticated_permisssion('result.publish_result')
class GetResults(View):

    model = Result

    def get(self, request):
        results = self.model.objects.all()
        return render(request, 'result/result_list.html', {'results': results})
 
@class_require_authenticated_permisssion('result.publish_result')
class CreateResult(View):

    model = Result
    form_class = ResultForm

    def get(self, request, student_slug, exam_slug):
        student = get_object_or_404(Student, slug = student_slug)
        exam = get_object_or_404(Exam, slug = exam_slug)       
        form =  self.form_class(initial={'student':student, 'exam':exam})
        return render(request, 'result/result_form.html', {'form':form, 'exam':exam, 'student':student})

    def post(self, request, student_slug, exam_slug):
        bound_form = self.form_class(request.POST)
        if bound_form.is_valid():
            newResult = bound_form.save(request)
            return redirect(newResult)
        else:
            return render(request, 'result/result.html', {'form':bound_form})
        
@class_require_authenticated_permisssion('result.publish_result')
class UpdateResult(View):

    model = Result
    form_class = ResultForm

    def get(self, request, result_slug):
        result = get_object_or_404(self.model, slug = result_slug)
        form = self.form_class(instance=result)
        return render(request, 'result/result_form_update.html', {'form':form, 'result': result})


    def post(self, request, result_slug):
        result = get_object_or_404(self.model, slug = result_slug)
        bound_form = self.form_class(request.POST, instance = result)
        if(bound_form.is_valid()):
            updatedExam = bound_form.save()
            messages.success(request, "Exam Updated")
            return redirect(updatedExam)
        return render(request, 'result/result_form_update.html' , {'form':bound_form, 'result': result})

@class_require_authenticated_permisssion('result.publish_result')
class DeleteResult(View):

    model = Result

    def get(self, request, result_slug):
        result = get_object_or_404(self.model, slug = result_slug)
        return render(request, 'result/result_confirm_delete.html', {'result': result})


    def post(self, request, result_slug):
        result = get_object_or_404(self.model, slug=result_slug)
        result.delete()
        return redirect('examination_exam_list')
    
    
@class_require_authenticated_permisssion('result.publish_result')
class DownloadResultFile(View):
    
    def get(self, request, exam_slug):
        
        exam = get_object_or_404(Exam, slug = exam_slug)
        
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f"attachment; filename={exam.examName}.xlsx"
        
        workbook = Workbook(response, {'in_memory': True})
        ws = workbook.add_worksheet()

        add_data_to_worksheet(workbook,ws, exam, Student)
        workbook.close()

        return response




